import csv
import io
import logging
from datetime import date

from dateutil import parser as dateutil_parser

from patients.models import ClassGroup, Department, Student, Employee

logger = logging.getLogger('imports.services')

_DEFAULT_CLASS_GROUP_ID = 'NAO_INFORMADO'
_DEFAULT_DEPARTMENT_ID  = 'NAO_INFORMADO'


# ──────────────────────────────────────────────────────────────────────────────
# Helpers de parsing
# ──────────────────────────────────────────────────────────────────────────────

def _parse_csv(file_obj) -> list:
    content = file_obj.read()
    if isinstance(content, bytes):
        content = content.decode('utf-8')
    reader = csv.DictReader(io.StringIO(content))
    return list(reader)


def _parse_xlsx(file_obj) -> list:
    import openpyxl
    wb = openpyxl.load_workbook(file_obj)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for raw_row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in raw_row):
            continue
        row_dict = {
            str(h): (str(v) if v is not None else '')
            for h, v in zip(headers, raw_row)
        }
        rows.append(row_dict)
    return rows


# ──────────────────────────────────────────────────────────────────────────────
# Helpers de cálculo e resolução de FK
# ──────────────────────────────────────────────────────────────────────────────

def _calculate_age(birth_date_str) -> int:
    if not birth_date_str or not str(birth_date_str).strip():
        return 0
    try:
        birth_date = dateutil_parser.parse(str(birth_date_str).strip(), dayfirst=True).date()
        today = date.today()
        return (
            today.year - birth_date.year
            - ((today.month, today.day) < (birth_date.month, birth_date.day))
        )
    except Exception:
        return 0


def _parse_birth_date(birth_date_str):
    if not birth_date_str or not str(birth_date_str).strip():
        return None
    try:
        from django.utils import timezone
        naive_dt = dateutil_parser.parse(str(birth_date_str).strip(), dayfirst=True)
        return timezone.make_aware(naive_dt) if timezone.is_naive(naive_dt) else naive_dt
    except Exception:
        return None


def _resolve_class_group(class_group_id) -> ClassGroup:
    cg_id = str(class_group_id).strip() if class_group_id else ''
    if cg_id:
        try:
            return ClassGroup.objects.get(id=cg_id)
        except ClassGroup.DoesNotExist:
            logger.warning(f"ClassGroup '{cg_id}' não encontrado. Usando padrão.")

    default, created = ClassGroup.objects.get_or_create(
        id=_DEFAULT_CLASS_GROUP_ID,
        defaults={
            'name': 'Turma não informada',
            'segment': 'Não informado',
            'director': 'Não informado',
        },
    )
    if created:
        logger.info("ClassGroup padrão 'NAO_INFORMADO' criado automaticamente.")
    return default


def _resolve_department(department_id) -> Department:
    dep_id = str(department_id).strip() if department_id else ''
    if dep_id:
        try:
            return Department.objects.get(id=dep_id)
        except Department.DoesNotExist:
            logger.warning(f"Department '{dep_id}' não encontrado. Usando padrão.")

    default, created = Department.objects.get_or_create(
        id=_DEFAULT_DEPARTMENT_ID,
        defaults={
            'name': 'Departamento não informado',
            'director': 'Não informado',
        },
    )
    if created:
        logger.info("Department padrão 'NAO_INFORMADO' criado automaticamente.")
    return default


# ──────────────────────────────────────────────────────────────────────────────
# Helpers de upsert
# ──────────────────────────────────────────────────────────────────────────────

def _upsert_student(row: dict, errors: list, row_num: int) -> str:
    registry = row.get('registry', '').strip()
    name     = row.get('name', '').strip()
    gender   = row.get('gender', '').strip()

    if not registry:
        errors.append({'row': row_num, 'reason': "Campo 'registry' ausente"})
        return 'error'
    if not name:
        errors.append({'row': row_num, 'reason': "Campo 'name' ausente"})
        return 'error'
    if not gender:
        errors.append({'row': row_num, 'reason': "Campo 'gender' ausente"})
        return 'error'

    birth_date_str = row.get('birth_date', '')
    age            = _calculate_age(birth_date_str)
    birth_date     = _parse_birth_date(birth_date_str)
    class_group    = _resolve_class_group(row.get('class_group_id', ''))

    email = row.get('email', '').strip() or None

    fields = {
        'name':         name,
        'gender':       gender,
        'age':          age,
        'email':        email,
        'birth_date':   birth_date,
        'class_group':  class_group,
        'father_name':  row.get('father_name', '').strip() or None,
        'father_phone': row.get('father_phone', '').strip() or None,
        'mother_name':  row.get('mother_name', '').strip() or None,
        'mother_phone': row.get('mother_phone', '').strip() or None,
    }

    try:
        student = Student.objects.get(registry=registry)
        for attr, value in fields.items():
            setattr(student, attr, value)
        student.save()
        logger.debug(f"Aluno atualizado: registry={registry}")
        return 'updated'
    except Student.DoesNotExist:
        Student.objects.create(id=registry, registry=registry, **fields)
        logger.debug(f"Aluno criado: registry={registry}")
        return 'created'


def _upsert_employee(row: dict, errors: list, row_num: int) -> str:
    registry = row.get('registry', '').strip()
    name     = row.get('name', '').strip()
    gender   = row.get('gender', '').strip()

    if not registry:
        errors.append({'row': row_num, 'reason': "Campo 'registry' ausente"})
        return 'error'
    if not name:
        errors.append({'row': row_num, 'reason': "Campo 'name' ausente"})
        return 'error'
    if not gender:
        errors.append({'row': row_num, 'reason': "Campo 'gender' ausente"})
        return 'error'

    birth_date_str = row.get('birth_date', '')
    age            = _calculate_age(birth_date_str)
    birth_date     = _parse_birth_date(birth_date_str)
    department     = _resolve_department(row.get('department_id', ''))

    email = row.get('email', '').strip() or None

    fields = {
        'name':       name,
        'gender':     gender,
        'age':        age,
        'email':      email,
        'birth_date': birth_date,
        'department': department,
        'position':   row.get('position', '').strip() or None,
    }

    try:
        employee = Employee.objects.get(registry=registry)
        for attr, value in fields.items():
            setattr(employee, attr, value)
        employee.save()
        logger.debug(f"Funcionário atualizado: registry={registry}")
        return 'updated'
    except Employee.DoesNotExist:
        Employee.objects.create(id=registry, registry=registry, **fields)
        logger.debug(f"Funcionário criado: registry={registry}")
        return 'created'


# ──────────────────────────────────────────────────────────────────────────────
# Função principal
# ──────────────────────────────────────────────────────────────────────────────

def import_from_file(file, entity_type: str, user):
    from .models import ImportLog

    valid_entities = {ImportLog.ENTITY_STUDENTS, ImportLog.ENTITY_EMPLOYEES}
    if entity_type not in valid_entities:
        raise ValueError(
            f"entity_type '{entity_type}' inválido. Use: {', '.join(valid_entities)}."
        )

    filename = getattr(file, 'name', '')
    name_lower = filename.lower()

    if name_lower.endswith('.csv'):
        rows = _parse_csv(file)
    elif name_lower.endswith('.xlsx'):
        rows = _parse_xlsx(file)
    else:
        raise ValueError(
            f"Formato de arquivo não suportado: '{filename}'. Use .csv ou .xlsx."
        )

    errors: list = []
    created_count = 0
    updated_count = 0

    upsert_fn = _upsert_student if entity_type == ImportLog.ENTITY_STUDENTS else _upsert_employee

    for row_num, row in enumerate(rows, start=1):
        result = upsert_fn(row, errors, row_num)
        if result == 'created':
            created_count += 1
        elif result == 'updated':
            updated_count += 1

    error_count = len(errors)

    log = ImportLog.objects.create(
        entity_type=entity_type,
        source=ImportLog.SOURCE_FILE,
        filename=filename,
        imported_by=user,
        total_rows=len(rows),
        created_count=created_count,
        updated_count=updated_count,
        error_count=error_count,
        errors=errors,
    )

    logger.info(
        f"Importação concluída — entidade={entity_type}, arquivo='{filename}', "
        f"total={len(rows)}, criados={created_count}, atualizados={updated_count}, "
        f"erros={error_count}"
    )
    return log
